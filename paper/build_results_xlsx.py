#!/usr/bin/env python3
"""Consolidate all OVLA experiment results (Final/evaluation/results/*.json) into
one .xlsx workbook, one sheet per experiment, with RAW per-item rows enriched by
the source dataset (command, ir_gt, connected_devices/selector context, has_else).
"""
import csv, json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = "/home/gnltnwjstk/joi"
R = os.path.join(ROOT, "paper/Final/evaluation/results")
OUT = os.path.join(R, "ovla_results.xlsx")

# ── dataset enrichment: authoritative name->dataset map ──
# result naming is NOT contiguous per category, so map via rq3's (name, cat, command)
# matched to dataset by (category_v2, command_eng); positional counter as fallback.
ds = list(csv.DictReader(open(os.path.join(ROOT, "dataset.csv"))))
ds_by_cmd = {}            # (cat, command_eng) -> row
counter = {}
pos_by_name = {}          # fallback positional
for row in ds:
    cat = row["category_v2"]
    ds_by_cmd[(cat, row["command_eng"].strip())] = row
    counter[cat] = counter.get(cat, 0) + 1
    pos_by_name[f"{cat}_{counter[cat]}"] = row

by_name = {}
for r in json.load(open(os.path.join(R, "rq3_pipeline_effect.json")))["rows"]:
    hit = ds_by_cmd.get((r["cat"], r["command"].strip()))
    by_name[r["name"]] = hit if hit else pos_by_name.get(r["name"], {})

def enrich(name):
    base = name.split("__")[0]  # strip mutation suffix
    d = by_name.get(base) or pos_by_name.get(base, {})
    return {
        "command": d.get("command_eng", ""),
        "ir_gt": d.get("ir_gt", ""),
        "connected_devices": d.get("connected_devices", ""),
        "has_else": d.get("has_else", ""),
        "notes": d.get("notes", ""),
    }

def load(f):
    return json.load(open(os.path.join(R, f)))

wb = Workbook()
wb.remove(wb.active)
HDR = Font(bold=True, color="FFFFFF")
HFILL = PatternFill("solid", fgColor="305496")

def sheet(title, headers, rows):
    ws = wb.create_sheet(title[:31])
    ws.append(headers)
    for c in ws[1]:
        c.font = HDR; c.fill = HFILL; c.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    ws.freeze_panes = "A2"
    from openpyxl.utils import get_column_letter
    for i, h in enumerate(headers, 1):
        L = get_column_letter(i)
        widths = [len(str(h))] + [len(str(r.get(h, ""))) for r in rows[:200]]
        ws.column_dimensions[L].width = min(max(max(widths) + 2, 10), 60)
    return ws

# ── 0. README ──
readme = [
    {"sheet": "README", "desc": "this index"},
    {"sheet": "rq3_safety", "desc": "RQ3 headline: 382 rows, verifier OFF vs ON, silent-wrong 9.2%->0"},
    {"sheet": "rq3_summary", "desc": "RQ3 aggregate (OFF/ON counts, reject breakdown)"},
    {"sheet": "nl2ir_errors", "desc": "382 rows: gen-IR vs gt-IR verdict (equiv / real error class)"},
    {"sheet": "instability", "desc": "600 rows (4 runs x 150 pairs): judge verdict flip on trace-identical variants"},
    {"sheet": "instability_summary", "desc": "flip rate by rewrite type / depth, per run"},
    {"sheet": "injected", "desc": "1038 rows (3 runs x 346): judge on injected-bug corpus (FN/FP)"},
    {"sheet": "injected_summary", "desc": "confusion matrix (TP/FP/TN/FN, recall, precision) per run"},
    {"sheet": "multigt", "desc": "60 rows (2 runs x 30): judge over-rejection on multi-GT equivalents"},
    {"sheet": "faithfulness", "desc": "RQ1: per-fault-class surface rate (Part B synthetic) + Part A real"},
]
sheet("README", ["sheet", "desc"], readme)

# ── 1. rq3_safety (382 raw) ──
d = load("rq3_pipeline_effect.json")
rows = []
for r in d["rows"]:
    e = enrich(r["name"])
    rows.append({**r,
                 "attempt_kinds": json.dumps(r.get("attempt_kinds"), ensure_ascii=False),
                 "final_violation_kinds": json.dumps(r.get("final_violation_kinds"), ensure_ascii=False),
                 "ir_gt": e["ir_gt"], "connected_devices": e["connected_devices"], "has_else": e["has_else"]})
sheet("rq3_safety",
      ["name", "cat", "command", "off_verdict", "on_verdict", "off_ok", "on_ok", "accepted",
       "n_attempts", "attempt_kinds", "final_violation_kinds", "off_msg",
       "has_else", "ir_gt", "connected_devices"], rows)

# rq3 summary
s = d["summary"]
srows = []
for phase in ("OFF", "ON"):
    for k, v in s[phase].items():
        srows.append({"phase": phase, "metric": k, "value": v})
srows.append({"phase": "headline", "metric": "headline", "value": s.get("headline", "")})
ra = s.get("reject_analysis", {})
for k, v in ra.get("by_category", {}).items():
    srows.append({"phase": "reject_by_category", "metric": k, "value": v})
for k, v in ra.get("by_final_violation_kind", {}).items():
    srows.append({"phase": "reject_by_violation", "metric": k, "value": v})
sheet("rq3_summary", ["phase", "metric", "value"], srows)

# ── 2. nl2ir_errors (382 raw) ──
d = load("nl2ir_error_distribution.json")
rows = []
for r in d["rows"]:
    e = enrich(r["name"])
    rows.append({**r, "command": e["command"], "ir_gt": e["ir_gt"], "has_else": e["has_else"]})
sheet("nl2ir_errors", ["name", "cat", "verdict", "reason", "command", "has_else", "ir_gt"], rows)

# ── 3. instability (4 runs x 150) ──
rows = []
for f, model, mode in [("instability_9B.json", "9B", "direct"),
                        ("instability_gpt51.json", "gpt-5.1", "direct"),
                        ("instability_9B_retrans.json", "9B", "roundtrip"),
                        ("instability_gpt51_retrans.json", "gpt-5.1", "roundtrip")]:
    d = load(f)
    for r in d["detail"]:
        e = enrich(r["seed"])
        rows.append({"model": model, "mode": mode, **r, "command": e["command"]})
sheet("instability",
      ["model", "mode", "seed", "type", "depth", "base_verdict", "variant_verdict", "flip", "command"], rows)

# instability summary
rows = []
for f, model, mode in [("instability_9B.json", "9B", "direct"),
                        ("instability_gpt51.json", "gpt-5.1", "direct"),
                        ("instability_9B_retrans.json", "9B", "roundtrip"),
                        ("instability_gpt51_retrans.json", "gpt-5.1", "roundtrip")]:
    d = load(f)
    rows.append({"model": model, "mode": mode, "group": "OVERALL", "flips": d["overall_flip"], "rate": d["overall_rate"]})
    for t, (fl, n) in d["by_type"].items():
        rows.append({"model": model, "mode": mode, "group": f"type:{t}", "flips": f"{fl}/{n}", "rate": round(fl / n, 4)})
    for t, (fl, n) in d["by_depth"].items():
        rows.append({"model": model, "mode": mode, "group": f"depth:{t}", "flips": f"{fl}/{n}", "rate": round(fl / n, 4)})
sheet("instability_summary", ["model", "mode", "group", "flips", "rate"], rows)

# ── 4. injected (3 runs x 346) ──
rows = []
for f, model, mode in [("injected_9B_direct.json", "9B", "direct"),
                       ("injected_gpt51_direct.json", "gpt-5.1", "direct"),
                       ("injected_9B_retrans.json", "9B", "roundtrip")]:
    d = load(f)
    for r in d["detail"]:
        e = enrich(r["name"])
        rows.append({"model": model, "mode": mode, "base_name": r["name"].split("__")[0], **r,
                     "command": e["command"]})
sheet("injected",
      ["model", "mode", "name", "base_name", "gt", "judge_says_wrong", "fault_family", "construct",
       "operator", "command", "problem"], rows)

# injected summary
rows = []
for f, model, mode in [("injected_9B_direct.json", "9B", "direct"),
                       ("injected_gpt51_direct.json", "gpt-5.1", "direct"),
                       ("injected_9B_retrans.json", "9B", "roundtrip")]:
    d = load(f); j = d["judge"]
    rows.append({"model": model, "mode": mode, **j})
sheet("injected_summary",
      ["model", "mode", "TP", "FP", "TN", "FN", "unparsed", "recall", "precision"], rows)

# ── 5. multigt (2 x 30) ──
rows = []
for f, model in [("multigt_9B.json", "9B"), ("multigt_gpt51_explicit.json", "gpt-5.1")]:
    d = load(f)
    for r in d["detail"]:
        e = enrich(r["name"])
        rows.append({"model": model, **r, "command": e["command"]})
sheet("multigt", ["model", "name", "gt", "judge_says_wrong", "command", "problem"], rows)

# ── 6. faithfulness (per class) ──
d = load("faithfulness_surfacing.json")
rows = []
pa = d["part_A_real_errors"]
rows.append({"part": "A_real", "class": "(all real NL->IR errors)",
             "applicable": pa["n"], "surfaced": pa["surfaced"],
             "surface_rate": pa["surface_rate"], "blind_spots": json.dumps(pa["blind_spots"], ensure_ascii=False)})
for cls, v in d["part_B_synthetic_by_class"].items():
    rows.append({"part": "B_synthetic", "class": cls,
                 "applicable": v.get("applicable"), "surfaced": v.get("surfaced"),
                 "surface_rate": round(v["surfaced"] / max(1, v["applicable"]), 4),
                 "blind_spots": json.dumps(v.get("blind_spots", []), ensure_ascii=False)})
sheet("faithfulness", ["part", "class", "applicable", "surfaced", "surface_rate", "blind_spots"], rows)

wb.save(OUT)
print("saved", OUT)
# coverage check for enrichment
miss = set()
for r in load("rq3_pipeline_effect.json")["rows"]:
    if r["name"] not in by_name:
        miss.add(r["name"])
print("rq3 names not matched to dataset:", len(miss), sorted(miss)[:10])
