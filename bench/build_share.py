#!/usr/bin/env python3
"""공유용 데이터셋 — 우리만 쓰는 열을 덜어내고 도메인별로 나눈다.

왜 따로 만드나
  dataset_5k.csv 는 만드는 쪽 열(어느 틀을 썼나, 어느 카테고리를 골랐나)까지 다 들고
  있다. 밖으로 나갈 때는 문제와 정답만 있으면 된다. 그리고 받는 쪽은 도메인별로
  나뉜 시트를 원한다.

  뺀 열 9개: mode trig act dev_trig dev_act ref b1 b3 ir_gt
  더한 열 2개: domain(집·사무실·연구실·농장·공장) intent(무엇을 시키는 말인가)

intent 는 새로 매기는 게 아니라 **빼는 열 셋에서 뽑아 둔 것**이다. act·b1·trig 가
사라지면 "기기를 시키는 말인지, 값을 묻는 말인지" 를 알 방법이 없어지기 때문이다.

    python bench/build_share.py          # → bench/share/*.csv + joi_bench.xlsx
"""
import collections
import csv
import json
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

import ir as IR            # noqa: E402
import korean as K         # noqa: E402
import policy as P         # noqa: E402

OUT = os.path.join(HERE, "share")

# 밖으로 안 내보내는 열 — 만드는 쪽 사정이다
DROP = ("mode", "trig", "act", "dev_trig", "dev_act", "ref", "b1", "b3", "ir_gt")

COLS = ["id", "domain", "space_id", "intent", "command", "command_ko",
        "expect", "why", "d", "tier", "context", "targets", "target_svc"]

KIND_KO = {"home": "집", "office": "사무실", "lab": "연구실",
           "farm": "농장", "factory": "공장"}

# 열이 무슨 뜻인지 — 받는 쪽이 읽을 안내다
GUIDE = [
    ("id", "G = 우리가 만든 5,000개(티어 오름차순), U = JoI use case 578개"),
    ("domain", "집·사무실·연구실·농장·공장. 시트가 이 값으로 나뉜다"),
    ("space_id", "어느 공간에서 말했나. spaces.json 의 열쇠"),
    ("intent", "무엇을 시키는 말인가 — 기기제어 / 알림예약 / 상태조회 / 외부정보"),
    ("command", "영어 문장 (원본)"),
    ("command_ko", "한국어 문장. 번역이 아니라 같은 자리에 한국어를 끼워 따로 만든 것"),
    ("expect", "정답 판정 — execute 실행 / ask 되묻기 / refuse 거절"),
    ("why", "거절 사유. no_device 기기 없음 / no_service 카탈로그에 그 서비스 없음 / "
            "no_channel 문장이 댄 통로(문자·슬랙·메일)가 없음 / "
            "no_occupancy 재실을 볼 방법 없음 / no_context 바깥 정보 못 읽음"),
    ("d", "시간·로직의 모양. D1 지금 한 번 · D4 무슨 일이 생기면 · D7 주기 · "
          "D10 제한시간 대기 · D11 두 번 읽고 비교 · D12 누적 · D13 복합"),
    ("tier", "난이도. T0 즉시 · T1 트리거 하나→동작 하나 · T2 조건·지연 · "
             "T3 반복·제한시간 · T4 변수·비교·누적"),
    ("context", "집 밖 정보가 필요한가 — none / phone / sun / weather / calendar"),
    ("targets", "정답 기기 id, 공백으로 나눔. expect=ask 면 정답이 아니라 선택지다"),
    ("target_svc", "정답 서비스. 알림처럼 기기가 아니라 채널을 고르는 명령에 붙는다"),
]

INTENT_GUIDE = [
    ("기기제어", "기기를 움직인다", "거실 조명 켜 줘 / 문 열리면 에어컨 꺼"),
    ("알림예약", "무슨 일이 생기면 알려 달라", "비 예보가 있으면 알려 줘"),
    ("상태조회", "지금 값을 묻는다", "지금 거실 온도 어때? / 현관문 열려 있어?"),
    ("외부정보", "날씨·일정을 묻는다", "오늘 비 와? / 다음 회의 몇 시야?"),
]

# ── 숫자 기준 (policy.py · ir.py) 을 안내 시트에 그대로 편다 ──────────────
# "너무 더우면 에어컨 켜 줘" 를 채점하려면 26℃ 라는 걸 알아야 한다. 문장 표만
# 넘기면 받는 쪽이 이 값을 모른다.
CONST_KO = {
    "too_warm_c":         "덥다 (℃)",
    "too_cold_c":         "춥다 (℃)",
    "too_humid_pct":      "눅눅하다 (%)",
    "too_dry_pct":        "건조하다 (%)",
    "dust_bad_ugm3":      "미세먼지 나쁨 PM10 (㎍/㎥)",
    "fine_dust_bad_ugm3": "초미세먼지 나쁨 PM2.5 (㎍/㎥)",
    "co2_high_ppm":       "이산화탄소 높다 (ppm)",
    "too_dark_lux":       "어둡다 (lux)",
    "too_bright_lux":     "눈부시다 (lux)",
    "too_loud_db":        "시끄럽다 (dB)",
    "battery_low_pct":    "배터리 부족 (%)",
    "tank_low_pct":       "탱크 부족 (%)",
    "soil_dry_pct":       "흙이 마름 (%)",
    "power_spike_w":      "전력 급증 (W)",
    "wind_strong_ms":     "바람 강함 (m/s)",
    "gas_danger_ppm":     "가스 위험 (ppm)",
    "vibration_high_mms": "진동 높음 (mm/s)",
}

DELTA_KO = [
    ("thermal_comfort", "체감 온도", "\"시원하게\" = 지금 온도 − 2℃"),
    ("temperature", "온도", "\"조금 따뜻하게\" = 지금 온도 + 2℃"),
    ("humidity", "습도", "\"눅눅해\" = 지금 습도 − 10%p"),
    ("illuminance", "밝기", "\"어둡게\" = 지금 밝기의 절반 (\"밝게\" 는 80% 로)"),
    ("sound", "소리", "볼륨 한 단계"),
]

CONST_NOTE = (
    "숫자 없는 말을 정답으로 바꾸는 표다. \"덥다\" 는 집에서 26℃, 온실에서 30℃ 다. "
    "사용자마다 다르게 두지 않고 공간 종류 다섯으로 고정했다 — 사용자 설정에 따라 "
    "정답이 갈리면 채점이 두 갈래가 되기 때문이다.\n"
    "근거: 온·습도는 국내 실내 권장(여름 26℃ / 겨울 20℃, 습도 40~60%), 미세먼지는 "
    "환경부 '나쁨' 경계(PM10 81, PM2.5 36), CO₂ 는 실내공기질 1,000ppm. "
    "농장·공장 값과 전력 기준은 근거 문헌이 아니라 우리가 정한 설정값이다."
)

SCENE_NOTE = (
    "카탈로그에 장면(Scene) 서비스가 없다. \"영화 모드\" 는 조명 값의 묶음으로 푼다 — "
    "밝기 한 번, 색온도 한 번, 호출 두 번이다."
)


def ir_ops(raw):
    """정답 IR 에 쓰인 op 이름 모음. IR 이 없으면 빈 집합."""
    if not raw:
        return set()
    out = set()

    def walk(nodes):
        for n in nodes:
            out.add(n.get("op"))
            for k in ("body", "then", "else"):
                if isinstance(n.get(k), list):
                    walk(n[k])
    try:
        walk(json.loads(raw)["timeline"])
    except Exception:                                   # noqa: BLE001
        return set()
    return out


# 집 밖 정보를 읽어 주는 기기 — 이걸 읽으면 "날씨·일정 묻기" 다
OUTSIDE = ("WeatherProvider", "CalendarProvider", "SunProvider")


def read_srcs(raw):
    """정답 IR 의 read 마디가 무엇을 읽는가."""
    if not raw:
        return []
    out = []

    def walk(nodes):
        for n in nodes:
            if n.get("op") == "read" and n.get("src"):
                out.append(n["src"])
            for k in ("body", "then", "else"):
                if isinstance(n.get(k), list):
                    walk(n[k])
    try:
        walk(json.loads(raw)["timeline"])
    except Exception:                                   # noqa: BLE001
        return []
    return out


def intent_of(r):
    """무엇을 시키는 말인가 — act·b1·trig·정답 IR 에서 뽑는다.

    까다로운 자리가 둘이다.

    하나. 알림(notify)이 두 갈래다. "거실 온도 보내 줘" 는 지금 값을 읽어 답하는
    조회고, "문 열리면 알려 줘" 는 나중을 예약하는 말이다. 문장만 보면 둘 다
    "알려 줘" 라서 **정답 IR 에 read 마디가 있는지**로 가른다.

    둘. 날씨 묻기와 "날씨가 바뀌면" 을 헷갈리면 안 된다. context 열은 **방아쇠가**
    바깥 정보를 쓰는지를 말할 뿐이라, 그걸로 가르면 "서리 예보가 있으면 전력 얼마나
    쓰고 있어" 가 날씨 질문이 되어 버린다. **read 마디가 무엇을 읽는지**로 가른다 —
    날씨·달력을 읽으면 바깥 정보를 묻는 말이고, 우리 센서를 읽으면 상태 조회다.

    되묻기·거절 행은 IR 이 없으므로 b1·trig·context 로 떨어진다.
    """
    if r["act"] in ("query", "notify"):
        # "@count:" 는 값을 묻는 읽기가 아니라 방아쇠가 몇 번 일어났는지 세는 것이다
        # (D12 "바쁜 날이 3번 넘으면"). 답할 값을 읽는 마디만 남긴다.
        srcs = [s for s in read_srcs(r["ir_gt"]) if "@count:" not in s]
        if srcs:
            # 바깥 정보를 "묻는" 것은 조회 문장뿐이다.
            if r["act"] == "query" and any(s.split(".")[0] in OUTSIDE for s in srcs):
                return "외부정보"
            return "상태조회"
        if r["ir_gt"]:                      # 실행인데 read 가 없다 = 그냥 알림
            return "알림예약" if r["act"] == "notify" else "상태조회"
        # 되묻기·거절 — 정답 IR 이 없다
        if r["act"] == "query":
            return ("외부정보"
                    if r["trig"] == "now" and r["context"] in ("weather", "calendar")
                    else "상태조회")
        return "상태조회" if "read" in r["b1"] else "알림예약"
    return "기기제어"


def main():
    rows = list(csv.DictReader(
        open(os.path.join(HERE, "dataset_5k.csv"), encoding="utf-8")))
    spaces = json.load(open(os.path.join(HERE, "spaces.json"), encoding="utf-8"))["spaces"]
    items = spaces.items() if isinstance(spaces, dict) else [(x["id"], x) for x in spaces]
    kind = {sid: s["kind"] for sid, s in items}
    name_ko = {sid: s.get("name_ko", "") for sid, s in items}

    out = []
    for r in rows:
        k = kind[r["space_id"]]
        row = {c: r.get(c, "") for c in COLS}
        row["domain"] = KIND_KO[k]
        row["intent"] = intent_of(r)
        out.append((k, row))

    os.makedirs(OUT, exist_ok=True)
    byk = collections.defaultdict(list)
    for k, row in out:
        byk[k].append(row)

    # ── CSV 5장 + 전체 1장
    for k in ("home", "office", "lab", "farm", "factory"):
        with open(os.path.join(OUT, f"joi_bench_{k}.csv"), "w",
                  encoding="utf-8-sig", newline="") as f:
            w = csv.DictWriter(f, COLS)
            w.writeheader()
            w.writerows(byk[k])
    with open(os.path.join(OUT, "joi_bench_all.csv"), "w",
              encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, COLS)
        w.writeheader()
        w.writerows(r for _, r in out)

    # ── 엑셀 한 권, 시트 7장 (안내 + 공간표 + 도메인 5)
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl 이 없어 xlsx 는 건너뛴다 (CSV 는 만들어졌다)")
        openpyxl = None

    if openpyxl:
        wb = openpyxl.Workbook()
        head_fill = PatternFill("solid", fgColor="EFEAEE")
        head_font = Font(bold=True)

        ws = wb.active
        ws.title = "안내"
        num_rows = []                       # 오른쪽 정렬할 자리 (행, 열들)

        def head(cells):
            """머리글 줄 — 굵게 + 바탕색."""
            r = ws.max_row
            for c in cells:
                ws[f"{c}{r}"].font = head_font
                ws[f"{c}{r}"].fill = head_fill

        def wide(text, cols="B:F"):
            """긴 글 한 줄 — 오른쪽 칸을 합쳐 넓게 쓴다."""
            ws.append([text] if cols == "A:F" else ["", text])
            r = ws.max_row
            a, b = cols.split(":")
            if cols == "A:F":
                ws.merge_cells(f"A{r}:F{r}")
            else:
                ws.merge_cells(f"{a}{r}:{b}{r}")
            return r

        ws.append(["JoI 명령어 벤치마크 — 공유용"])
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:F1")
        ws.append([])
        wide(f"명령어 {len(out):,}개. 영어와 한국어가 같은 줄에 있다. "
             "공간 40곳의 기기 목록이 따로 있어서, 같은 문장이라도 "
             "어디서 말했느냐에 따라 실행·되묻기·거절로 갈린다.", "A:F")
        wide("도메인마다 시트를 나눴다. 만드는 쪽에서만 쓰는 열 9개"
             "(mode trig act dev_trig dev_act ref b1 b3 ir_gt)는 뺐다.", "A:F")

        # ── 열 사전
        ws.append([])
        ws.append(["열", "뜻"])
        ws.merge_cells(f"B{ws.max_row}:F{ws.max_row}")
        head("AB")
        for name, desc in GUIDE:
            ws.append([name, desc])
            ws.merge_cells(f"B{ws.max_row}:F{ws.max_row}")

        # ── intent
        ws.append([])
        ws.append(["intent", "뜻", "", "예"])
        r = ws.max_row
        ws.merge_cells(f"B{r}:C{r}")
        ws.merge_cells(f"D{r}:F{r}")
        head("ABD")
        for a, b, c in INTENT_GUIDE:
            ws.append([a, b, "", c])
            r = ws.max_row
            ws.merge_cells(f"B{r}:C{r}")
            ws.merge_cells(f"D{r}:F{r}")

        # ── 기준값 — "덥다" 가 몇 도인가
        ws.append([])
        ws.append(["숫자 기준 — 숫자 없는 말을 정답으로 바꾸는 표"])
        ws[f"A{ws.max_row}"].font = Font(bold=True, size=12)
        ws.merge_cells(f"A{ws.max_row}:F{ws.max_row}")
        wide(CONST_NOTE, "A:F")
        ws.append([])
        ws.append(["기준"] + [KIND_KO[k] for k in P.KINDS])
        head("ABCDEF")
        for name, vals in P.CONST.items():
            ws.append([CONST_KO.get(name, name), *vals])
            num_rows.append((ws.max_row, "BCDEF"))

        # ── 값 보폭 — "조금 더" 가 얼마인가
        ws.append([])
        ws.append(["\"조금 더\" 가 얼마인가", "보폭", "", "예"])
        r = ws.max_row
        ws.merge_cells(f"B{r}:C{r}")
        ws.merge_cells(f"D{r}:F{r}")
        head("ABD")
        step_ko = {"illuminance": "×0.5", "humidity": "10%p",
                   "thermal_comfort": "2℃", "temperature": "2℃"}
        for key, ko, ex in DELTA_KO:
            ws.append([ko, step_ko.get(key, P.DELTA[key]), "", ex])
            r = ws.max_row
            ws.merge_cells(f"B{r}:C{r}")
            ws.merge_cells(f"D{r}:F{r}")
            num_rows.append((r, "B"))

        # ── 장면 — 카탈로그에 없어서 조명 값으로 푼다
        ws.append([])
        ws.append(["장면 — 조명 값의 묶음으로 푼다"])
        ws[f"A{ws.max_row}"].font = Font(bold=True, size=12)
        ws.merge_cells(f"A{ws.max_row}:F{ws.max_row}")
        wide(SCENE_NOTE, "A:F")
        ws.append([])
        ws.append(["장면", "가리키는 말", "밝기 %", "색온도 K", "색상 Hue", "채도 %"])
        head("ABCDEF")
        for name, parts in IR.SCENE.items():
            d = dict(parts)
            ws.append([name, K.SCENE_KO.get(name, ""),
                       d.get("bri", "끈다" if "off" in d else ""),
                       d.get("k", ""),
                       IR.HUE[d["hue"]] if "hue" in d else "",
                       100 if "hue" in d else ""])
            num_rows.append((ws.max_row, "CDEF"))

        for col, wdt in zip("ABCDEF", (30, 15, 13, 13, 13, 13)):
            ws.column_dimensions[col].width = wdt
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for r, cols in num_rows:            # 숫자는 오른쪽으로
            for c in cols:
                ws[f"{c}{r}"].alignment = Alignment(vertical="top",
                                                    horizontal="right")

        ws = wb.create_sheet("공간")
        ws.append(["space_id", "domain", "어떤 곳", "기기 수", "문장 수"])
        for c in "ABCDE":
            ws[f"{c}1"].font = head_font
            ws[f"{c}1"].fill = head_fill
        n_sent = collections.Counter(r["space_id"] for _, r in out)
        for sid, s in sorted(items):
            ws.append([sid, KIND_KO[s["kind"]], name_ko.get(sid, ""),
                       len(s.get("devices", {})), n_sent[sid]])
        for col, wdt in zip("ABCDE", (12, 10, 32, 10, 10)):
            ws.column_dimensions[col].width = wdt
        ws.freeze_panes = "A2"

        for k in ("home", "office", "lab", "farm", "factory"):
            ws = wb.create_sheet(KIND_KO[k])
            ws.append(COLS)
            for i, c in enumerate(COLS, 1):
                cell = ws.cell(1, i)
                cell.font = head_font
                cell.fill = head_fill
            for row in byk[k]:
                ws.append([row[c] for c in COLS])
            widths = {"id": 9, "domain": 8, "space_id": 11, "intent": 10,
                      "command": 52, "command_ko": 40, "expect": 9, "why": 14,
                      "d": 5, "tier": 5, "context": 9, "targets": 44, "target_svc": 24}
            for i, c in enumerate(COLS, 1):
                ws.column_dimensions[get_column_letter(i)].width = widths[c]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{ws.max_row}"
        wb.save(os.path.join(OUT, "joi_bench.xlsx"))

    # ── 보고
    kinds = ("home", "office", "lab", "farm", "factory")
    cats = ("기기제어", "알림예약", "상태조회", "외부정보")
    tbl = collections.defaultdict(collections.Counter)
    for k, row in out:
        tbl[k][row["intent"]] += 1
    print(f"공유용 {len(out):,}행 · 열 {len(COLS)}개 (뺀 열 {len(DROP)}개) → {OUT}")
    print("\n  도메인 ×  intent" + "".join(c.rjust(10) for c in cats) + "        계")
    for k in kinds:
        print(f"  {KIND_KO[k]:6s} {k:9s}"
              + "".join(str(tbl[k][c]).rjust(10) for c in cats)
              + str(sum(tbl[k].values())).rjust(11))
    print("  " + " " * 16
          + "".join(str(sum(tbl[k][c] for k in kinds)).rjust(10) for c in cats)
          + str(len(out)).rjust(11))

    bad = []
    for _, row in out:
        for c in DROP:
            if c in row:
                bad.append(f"{row['id']} 에 {c} 가 남음")
    if len(set(r["id"] for _, r in out)) != len(out):
        bad.append("id 가 겹친다")
    if sum(len(v) for v in byk.values()) != len(out):
        bad.append("도메인 나눔에서 행이 새거나 늘었다")
    print("\n  검산:", *bad, sep="\n    ") if bad else print("\n  검산: 어긋난 것 없음 ✅")
    return 1 if bad else 0


if __name__ == "__main__":
    sys.exit(main())
